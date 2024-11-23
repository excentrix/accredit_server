# your_app/utils/excel_styles.py
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

class ExcelStyles:
    @staticmethod
    def get_header_style():
        return {
            'font': Font(bold=True, size=12),
            'fill': PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'alignment': Alignment(wrap_text=True)
        }

    @staticmethod
    def get_subheader_style():
        return {
            'font': Font(bold=True, size=11),
            'fill': PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'alignment': Alignment(wrap_text=True)
        }

    @staticmethod
    def get_data_style():
        return {
            'font': Font(size=10),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'alignment': Alignment(wrap_text=True)
        }

    @staticmethod
    def get_title_style():
        return {
            'font': Font(bold=True, size=14),
            'alignment': Alignment(horizontal='left')
        }

    @staticmethod
    def apply_styles(cell, styles):
        for key, value in styles.items():
            setattr(cell, key, value)