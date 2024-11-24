# core/utils/excel_export.py
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from .excel_styles import ExcelStyles

class ExcelExporter:
    def __init__(self, template, academic_year):
        self.template = template
        self.academic_year = academic_year
        self.wb = Workbook()
        self.ws = self.wb.active
        self.current_row = 1
        
        # Use ExcelStyles instead of defining styles directly
        self.header_style = ExcelStyles.get_header_style()
        self.subheader_style = ExcelStyles.get_subheader_style()
        self.data_style = ExcelStyles.get_data_style()
        self.title_style = ExcelStyles.get_title_style()

    def _apply_styles(self, cell, styles):
        ExcelStyles.apply_styles(cell, styles)

    def _write_title_info(self):
        title_info = [
            f"Template: {self.template.name}",
            f"Code: {self.template.code}",
            f"Academic Year: {self.academic_year.name}",
            f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]

        for info in title_info:
            cell = self.ws.cell(row=self.current_row, column=1, value=info)
            ExcelStyles.apply_styles(cell, self.title_style)  # Use title style
            self.current_row += 1

        # Add spacing
        self.current_row += 1

    def _get_flattened_columns(self, columns):
        """Flatten nested column structure"""
        flat_columns = []
        for column in columns:
            if column['type'] == 'single':
                flat_columns.append(column)
            elif column['type'] == 'group':
                for nested_column in column['columns']:
                    # Create a new column object to avoid modifying the original
                    flattened_column = nested_column.copy()
                    # Construct the flattened name to match the data structure
                    flattened_column['name'] = f"{column['name']}_{nested_column['name']}"
                    # Also store the display name for better readability
                    flattened_column['display_name'] = f"{column.get('display_name', column['name'])} - {nested_column.get('display_name', nested_column['name'])}"
                    flat_columns.append(flattened_column)
        return flat_columns

    def _write_section(self, section_index, submissions):
        try:
            if not self.template.metadata or section_index >= len(self.template.metadata):
                print(f"Invalid section index: {section_index}")
                return

            section = self.template.metadata[section_index]
            
            # Write section headers if present
            if 'headers' in section and section['headers']:
                total_columns = sum(
                    len(col['columns']) if col['type'] == 'group' else 1 
                    for col in section['columns']
                )
                
                for header in section['headers']:
                    merge_range = f'A{self.current_row}:{get_column_letter(total_columns)}{self.current_row}'
                    self.ws.merge_cells(merge_range)
                    header_cell = self.ws.cell(row=self.current_row, column=1, value=header)
                    self._apply_styles(header_cell, self.header_style)
                    self.current_row += 1

            # Write column group headers and subheaders
            current_col = 1
            column_mapping = []  # Store column name mapping for data rows

            # First row: Group headers
            group_header_row = self.current_row
            for column in section['columns']:
                if column['type'] == 'group':
                    # Calculate span for group
                    colspan = len(column['columns'])
                    if colspan > 1:
                        merge_range = f'{get_column_letter(current_col)}{group_header_row}:{get_column_letter(current_col + colspan - 1)}{group_header_row}'
                        self.ws.merge_cells(merge_range)
                    
                    cell = self.ws.cell(
                        row=group_header_row,
                        column=current_col,
                        value=column.get('display_name', column['name'])
                    )
                    self._apply_styles(cell, self.header_style)

                    # Write subheaders in next row
                    for idx, subcol in enumerate(column['columns']):
                        subcol_name = f"{column['name']}_{subcol['name']}"
                        column_mapping.append(subcol_name)
                        
                        cell = self.ws.cell(
                            row=group_header_row + 1,
                            column=current_col + idx,
                            value=subcol.get('display_name', subcol['name'])
                        )
                        self._apply_styles(cell, self.subheader_style)
                        
                        # Set column width
                        col_letter = get_column_letter(current_col + idx)
                        self.ws.column_dimensions[col_letter].width = max(
                            len(str(subcol.get('display_name', subcol['name']))) + 2,
                            15
                        )
                    
                    current_col += colspan
                else:
                    # Single column
                    column_mapping.append(column['name'])
                    cell = self.ws.cell(
                        row=group_header_row,
                        column=current_col,
                        value=column.get('display_name', column['name'])
                    )
                    self._apply_styles(cell, self.header_style)
                    
                    # Set column width
                    col_letter = get_column_letter(current_col)
                    self.ws.column_dimensions[col_letter].width = max(
                        len(str(column.get('display_name', column['name']))) + 2,
                        15
                    )
                    
                    current_col += 1

            # Move to data rows
            self.current_row = group_header_row + 2

            # Write data rows
            for submission in submissions:
                rows = submission.data_rows.filter(section_index=section_index)
                for row_data in rows:
                    current_col = 1
                    data = row_data.data.get('data', row_data.data)  # Handle both nested and flat data
                    
                    for col_name in column_mapping:
                        value = data.get(col_name, '')
                        cell = self.ws.cell(
                            row=self.current_row,
                            column=current_col,
                            value=value
                        )
                        style = self.data_style.copy()
                        if len(str(value)) > 50:
                            style['alignment'] = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        self._apply_styles(cell, style)
                        current_col += 1
                    
                    self.current_row += 1

            # Add spacing after section
            self.current_row += 1

        except Exception as e:
            print(f"Error in _write_section: {str(e)}")
            raise

    def export_to_worksheet(self, ws, submissions):
        """Export to an existing worksheet"""
        try:
            self.ws = ws
            self.current_row = 1
            
            if not submissions.exists():
                return False
                
            # Write title information
            self._write_title_info()

            # Process each section
            if self.template.metadata:
                for section_index in range(len(self.template.metadata)):
                    self._write_section(section_index, submissions)
            else:
                print(f"No metadata found for template {self.template.code}")

            # Auto-adjust row heights
            for row in self.ws.rows:
                max_length = 0
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value).split('\n')))
                if max_length > 1:
                    self.ws.row_dimensions[cell.row].height = max_length * 15

            return True

        except Exception as e:
            print(f"Error in export_to_worksheet: {str(e)}")
            raise

    def export(self, submissions):
        if not submissions.exists():
            return self.wb

        self.ws.title = f"{self.template.code} Data"
        self.export_to_worksheet(self.ws, submissions)
        return self.wb