# core/tests/test_exports.py
import pytest
from django.urls import reverse
from rest_framework import status
from openpyxl import load_workbook
import io
from core.models import (
    User, Department, AcademicYear, Template, 
    DataSubmission, SubmissionData
)
from core.tests.test_data import (
    generate_programme_data, generate_course_data,
    generate_value_added_course_data
)

@pytest.mark.django_db
class TestExcelExport:
    def test_export_template_1_1(self, client, template_1_1):
        # Create test data
        academic_year = AcademicYear.objects.create(
            year="2023-2024",
            is_current=True
        )
        
        department = Department.objects.create(
            name="Computer Science",
            code="CS"
        )
        
        iqac_director = User.objects.create_user(
            username="director",
            password="test123",
            role="iqac_director"
        )
        
        # Create submission
        submission = DataSubmission.objects.create(
            department=department,
            academic_year=academic_year,
            template=template_1_1,
            submitted_by=iqac_director,
            verified_by=iqac_director,
            status='approved'
        )
        
        # Add test data rows
        for _ in range(5):
            prog_code, prog_name = generate_programme_data()
            SubmissionData.objects.create(
                submission=submission,
                row_number=_ + 1,
                data={
                    "programme_code": prog_code,
                    "programme_name": prog_name
                }
            )
        
        # Login as IQAC director
        client.force_login(iqac_director)
        
        # Make export request
        response = client.get(f"/api/export/?template_code=1.1&academic_year={academic_year.id}")
        
        assert response.status_code == status.HTTP_200_OK
        
        # Load the exported Excel file
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Verify headers
        assert ws['A1'].value == "1.1. Number of programmes offered during the year"
        assert ws['A2'].value == "Programme Code"
        assert ws['B2'].value == "Programme Name"
        
        # Verify data is present
        assert ws['A3'].value is not None
        assert ws['B3'].value is not None

    def test_export_template_1_1_3(self, client, template_1_1_3):
        # Create test data
        academic_year = AcademicYear.objects.create(
            year="2023-2024",
            is_current=True
        )
        
        departments = [
            Department.objects.create(name="Computer Science", code="CS"),
            Department.objects.create(name="Electrical Engineering", code="EE")
        ]
        
        iqac_director = User.objects.create_user(
            username="director",
            password="test123",
            role="iqac_director"
        )
        
        # Create submissions for multiple departments
        for dept in departments:
            submission = DataSubmission.objects.create(
                department=dept,
                academic_year=academic_year,
                template=template_1_1_3,
                submitted_by=iqac_director,
                verified_by=iqac_director,
                status='approved'
            )
            
            # Add test data rows
            for i in range(3):
                course_data = generate_course_data()
                SubmissionData.objects.create(
                    submission=submission,
                    row_number=i + 1,
                    data={
                        "course_name": course_data['name'],
                        "course_code": course_data['code'],
                        "activities": course_data['activities'],
                        "document_link": f"https://example.com/docs/{course_data['code']}"
                    }
                )
        
        # Login and export
        client.force_login(iqac_director)
        response = client.get(f"/api/export/?template_code=1.1.3&academic_year={academic_year.id}")
        
        assert response.status_code == status.HTTP_200_OK
        
        # Verify Excel content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Verify headers
        assert "1.1.3 Details of courses" in ws['A1'].value
        assert ws['A3'].value == "Name of the Course"
        assert ws['B3'].value == "Course Code"
        
        # Verify data from both departments is present
        assert ws['A4'].value is not None
        assert ws['B4'].value is not None

    def test_export_permissions(self, client, template_1_1):
        # Create test data
        academic_year = AcademicYear.objects.create(
            year="2023-2024",
            is_current=True
        )
        
        faculty_user = User.objects.create_user(
            username="faculty",
            password="test123",
            role="faculty"
        )
        
        client.force_login(faculty_user)
        response = client.get(f"/api/export/?template_code=1.1&academic_year={academic_year.id}")
        
        assert response.status_code == status.HTTP_403_FORBIDDEN


    def test_export_invalid_template(self, client):
        """Test export with non-existent template code"""
        academic_year = AcademicYear.objects.create(year="2023-2024")
        iqac_director = User.objects.create_user(
            username="director",
            password="test123",
            role="iqac_director"
        )
        
        client.force_login(iqac_director)
        response = client.get(f"/api/export/?template_code=invalid&academic_year={academic_year.id}")
        
        assert response.status_code == status.HTTP_404_NOT_FOUND
        assert "Invalid template_code" in response.json().get('error')

    def test_export_missing_parameters(self, client):
        """Test export with missing required parameters"""
        iqac_director = User.objects.create_user(
            username="director",
            password="test123",
            role="iqac_director"
        )
        
        client.force_login(iqac_director)
        response = client.get("/api/export/")
        
        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "required" in response.json().get('error')

    def test_export_no_data(self, client, template_1_1):
        """Test export when no submissions exist"""
        academic_year = AcademicYear.objects.create(year="2023-2024")
        iqac_director = User.objects.create_user(
            username="director",
            password="test123",
            role="iqac_director"
        )
        
        client.force_login(iqac_director)
        response = client.get(f"/api/export/?template_code=1.1&academic_year={academic_year.id}")
        
        assert response.status_code == status.HTTP_200_OK
        
        # Verify empty Excel file structure
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Headers should still be present
        assert "1.1" in ws['A1'].value
        # But no data rows
        assert ws['A3'].value is None

    # core/tests/test_exports.py

    def test_export_excel_formatting(self, client, template_1_1_3):
        """Test Excel file formatting details"""
        academic_year = AcademicYear.objects.create(year="2023-2024")
        department = Department.objects.create(name="Computer Science", code="CS")
        iqac_director = User.objects.create_user(
            username="director",
            password="test123",
            role="iqac_director"
        )
        
        submission = DataSubmission.objects.create(
            department=department,
            academic_year=academic_year,
            template=template_1_1_3,
            submitted_by=iqac_director,
            verified_by=iqac_director,
            status='approved'
        )
        
        # Add test data
        course_data = generate_course_data()
        SubmissionData.objects.create(
            submission=submission,
            row_number=1,
            data={
                "course_name": course_data['name'],
                "course_code": course_data['code'],
                "activities": course_data['activities'],
                "document_link": f"https://example.com/docs/{course_data['code']}"
            }
        )
        
        client.force_login(iqac_director)
        response = client.get(f"/api/export/?template_code=1.1.3&academic_year={academic_year.id}")
        
        assert response.status_code == status.HTTP_200_OK
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check if cells are merged
        merged_ranges = [str(merged_range) for merged_range in ws.merged_cells.ranges]
        assert any('A1:D1' in merged_range for merged_range in merged_ranges)
        
        # Check column widths
        assert ws.column_dimensions['A'].width >= 15  # Course name column should be wide
        assert ws.column_dimensions['C'].width >= 30  # Activities column should be very wide
        
        # Check text wrapping in activities cell
        cell = ws.cell(row=3, column=3)  # First data row, activities column
        assert cell.alignment is not None
        assert cell.alignment.wrap_text is True
        
        # Check header formatting
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        assert header_cell.alignment.horizontal == 'center'

    def test_export_multiple_departments_order(self, client, template_1_1):
        """Test that data from multiple departments is properly ordered"""
        academic_year = AcademicYear.objects.create(year="2023-2024")
        departments = [
            Department.objects.create(name="Computer Science", code="CS"),
            Department.objects.create(name="Electronics", code="EC"),
            Department.objects.create(name="Mechanical", code="ME")
        ]
        iqac_director = User.objects.create_user(
            username="director",
            password="test123",
            role="iqac_director"
        )
        
        # Create submissions in reverse order
        for dept in reversed(departments):
            submission = DataSubmission.objects.create(
                department=dept,
                academic_year=academic_year,
                template=template_1_1,
                submitted_by=iqac_director,
                verified_by=iqac_director,
                status='approved'
            )
            
            prog_code, prog_name = generate_programme_data()
            SubmissionData.objects.create(
                submission=submission,
                row_number=1,
                data={
                    "programme_code": f"{dept.code}_{prog_code}",
                    "programme_name": prog_name
                }
            )
        
        client.force_login(iqac_director)
        response = client.get(f"/api/export/?template_code=1.1&academic_year={academic_year.id}")
        
        assert response.status_code == status.HTTP_200_OK
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Get department codes from the first column of data rows
        dept_codes = []
        for row in range(3, 6):  # Rows 3 to 5
            cell = ws.cell(row=row, column=1)
            if cell.value:  # Check if cell has value
                code = cell.value.split('_')[0]  # Extract department code
                dept_codes.append(code)
        
        # Verify codes are in alphabetical order
        expected_codes = ['CS', 'EC', 'ME']
        assert dept_codes == expected_codes, f"Expected {expected_codes}, got {dept_codes}"
        
        
# core/tests/test_exports.py
import os
from datetime import datetime

@pytest.mark.django_db
class TestExcelGeneration:
    def test_generate_all_templates(self, client):
        """
        Generate sample Excel files for all templates with realistic data.
        Files will be saved in a 'test_exports' directory.
        """
        # Create directory for test exports
        export_dir = "test_exports"
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)

        # Create common test data
        academic_year = AcademicYear.objects.create(year="2023-2024")
        departments = [
            Department.objects.create(name="Computer Science", code="CS"),
            Department.objects.create(name="Electronics", code="EC"),
            Department.objects.create(name="Mechanical", code="ME")
        ]
        iqac_director = User.objects.create_user(
            username="director", password="test123", role="iqac_director"
        )

        # Template 1.1 - Programme list
        template_1_1 = Template.objects.create(
            code="1.1",
            name="Number of programmes offered during the year",
            headers=["1.1. Number of programmes offered during the year"],
            columns=[
                {"name": "programme_code", "display_name": "Programme Code", "type": "string"},
                {"name": "programme_name", "display_name": "Programme Name", "type": "string"}
            ]
        )

        # Template 1.1.2/1.2.2 - Programme revision details
        template_1_1_2 = Template.objects.create(
            code="1.1.2",
            name="Details of Programmes with syllabus revision",
            headers=[
                "1.1.2 Details of Programmes where syllabus revision was carried out during the year",
                "1.2.2 Details of Programmes offered through Choice Based Credit System (CBCS)/Elective Course System"
            ],
            columns=[
                {"name": "programme_code", "display_name": "Programme Code", "type": "string"},
                {"name": "programme_name", "display_name": "Programme Name", "type": "string"},
                {"name": "year_of_introduction", "display_name": "Year of introduction (Date)", "type": "string"},
                {"name": "cbcs_status", "display_name": "Status of implemetation of CBCS / Elective Course System (Yes/No)", "type": "string"},
                {"name": "cbcs_year", "display_name": "Year of implemetation of CBCS / Elective Course System", "type": "string"},
                {"name": "revision_year", "display_name": "Year of revision, if any", "type": "string"},
                {"name": "content_change", "display_name": "If revision has been carried out in the syllabus during the year, percentage of content added or replaced", "type": "string"},
                {"name": "document_link", "display_name": "Link to the relevant document", "type": "url"}
            ]
        )

        # Template 1.1.3/1.2.1 - Course details
        template_1_1_3 = Template.objects.create(
            code="1.1.3",
            name="Details of courses offered",
            headers=[
                "1.1.3 Details of courses offered by the institution that focus on employability/ entrepreneurship/ skill development during the year.",
                "1.2.1 Details of courses introduced across all programmes offered during the year"
            ],
            columns=[
                {"name": "course_name", "display_name": "Name of the Course", "type": "string"},
                {"name": "course_code", "display_name": "Course Code", "type": "string"},
                {"name": "activities", "display_name": "Activities/Content with a direct bearing on Employability/ Entrepreneurship/ Skill development", "type": "text"},
                {"name": "document_link", "display_name": "Link to the relevant document", "type": "url"}
            ]
        )

        # Template 1.3.2/1.3.3 - Value added courses
        template_1_3_2 = Template.objects.create(
            code="1.3.2",
            name="Details of value-added courses",
            headers=[
                "1.3.2 Details of value-added courses for imparting transferable and life skills offered during the year",
                "1.3.3 Number of students enrolled in the courses under 1.3.2"
            ],
            columns=[
                {"name": "course_name", "display_name": "Name of the value-added courses (with 30 or more contact hours) offered", "type": "string"},
                {"name": "course_code", "display_name": "Course Code, if any", "type": "string"},
                {"name": "times_offered", "display_name": "No. of times offered during the year", "type": "number"},
                {"name": "duration", "display_name": "Duration of course (in hours)", "type": "number"},
                {"name": "enrolled", "display_name": "Number of students enrolled during the year", "type": "number"},
                {"name": "completed", "display_name": "Number of students who completed the course during the year", "type": "number"}
            ]
        )

        templates = [template_1_1, template_1_1_2, template_1_1_3, template_1_3_2]
        
        # Generate sample data for each template
        for template in templates:
            for dept in departments:
                submission = DataSubmission.objects.create(
                    department=dept,
                    academic_year=academic_year,
                    template=template,
                    submitted_by=iqac_director,
                    verified_by=iqac_director,
                    status='approved'
                )

                # Generate appropriate data based on template
                if template.code == "1.1":
                    # Programme list data
                    programmes = [
                        ("BCA", "Bachelor of Computer Applications"),
                        ("MCA", "Master of Computer Applications"),
                        ("BTech", "Bachelor of Technology")
                    ]
                    for i, (prog_code, prog_name) in enumerate(programmes, 1):
                        SubmissionData.objects.create(
                            submission=submission,
                            row_number=i,
                            data={
                                "programme_code": f"{dept.code}_{prog_code}",
                                "programme_name": f"{prog_name} ({dept.name})"
                            }
                        )

                elif template.code == "1.1.2":
                    # Programme revision data
                    years = ["2020", "2021", "2022"]
                    for i, year in enumerate(years, 1):
                        SubmissionData.objects.create(
                            submission=submission,
                            row_number=i,
                            data={
                                "programme_code": f"{dept.code}_PROG{i}",
                                "programme_name": f"Programme {i} ({dept.name})",
                                "year_of_introduction": str(int(year) - 5),
                                "cbcs_status": "Yes",
                                "cbcs_year": year,
                                "revision_year": "2023",
                                "content_change": f"{20 * i}%",
                                "document_link": f"https://example.com/docs/{dept.code}/prog{i}"
                            }
                        )

                elif template.code == "1.1.3":
                    # Course details data
                    courses = [
                        {
                            "name": "Advanced Programming",
                            "code": "CS101",
                            "activities": "Python, Java, Web Development\nProject Management\nTeam Collaboration"
                        },
                        {
                            "name": "Data Science",
                            "code": "DS201",
                            "activities": "Statistical Analysis\nMachine Learning\nData Visualization"
                        },
                        {
                            "name": "Cloud Computing",
                            "code": "CC301",
                            "activities": "AWS Services\nCloud Architecture\nServerless Computing"
                        }
                    ]
                    for i, course in enumerate(courses, 1):
                        SubmissionData.objects.create(
                            submission=submission,
                            row_number=i,
                            data={
                                "course_name": f"{course['name']} ({dept.name})",
                                "course_code": f"{dept.code}_{course['code']}",
                                "activities": course['activities'],
                                "document_link": f"https://example.com/docs/{dept.code}/{course['code']}"
                            }
                        )

                elif template.code == "1.3.2":
                    # Value added courses data
                    courses = [
                        {
                            "name": "Professional Communication",
                            "duration": 45,
                            "enrolled": 100
                        },
                        {
                            "name": "Digital Marketing",
                            "duration": 60,
                            "enrolled": 80
                        },
                        {
                            "name": "Leadership Skills",
                            "duration": 40,
                            "enrolled": 120
                        }
                    ]
                    for i, course in enumerate(courses, 1):
                        completed = int(course['enrolled'] * 0.85)  # 85% completion rate
                        SubmissionData.objects.create(
                            submission=submission,
                            row_number=i,
                            data={
                                "course_name": f"{course['name']} ({dept.name})",
                                "course_code": f"{dept.code}_VAC{i}",
                                "times_offered": "2",
                                "duration": str(course['duration']),
                                "enrolled": str(course['enrolled']),
                                "completed": str(completed)
                            }
                        )

        # Export all templates
        client.force_login(iqac_director)
        
        for template in templates:
            response = client.get(f"/api/export/?template_code={template.code}&academic_year={academic_year.id}")
            assert response.status_code == status.HTTP_200_OK
            
            # Save the file
            filename = f"{template.code}_{academic_year.year}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            with open(filepath, 'wb') as f:
                f.write(response.content)
            
            print(f"Generated {filepath}")