from rest_framework import viewsets, status, permissions
from rest_framework.filters import OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.exceptions import NotFound
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.http import HttpResponse


from .services import AcademicYearTransitionService
from .tasks import process_academic_year_transition

from .filters import DataSubmissionFilter
from .utils.excel_export import ExcelExporter

from django.utils import timezone
import io
import re
from rest_framework.views import APIView

from django.db import models

from user_management.serializers import UserSerializer, DepartmentSerializer
from user_management.models import CustomUser as User, Department

from .models import (
    AcademicYearTransition, Criteria, SubmissionHistory, AcademicYear, Template, 
    DataSubmission, SubmissionData, Board, Template, DataSubmission
)

from rest_framework.viewsets import ViewSet
from .services import DashboardService


from .serializers import (
    CriteriaSerializer, AcademicYearSerializer,
    TemplateSerializer, DataSubmissionSerializer, SubmissionDataSerializer, BoardSerializer,
)

import openpyxl
from openpyxl import Workbook
from django.core.exceptions import ValidationError


from .permissions import IsFaculty, IsIQACDirector, IsAdmin
from user_management.permissions import HasPermission


import logging
from deepdiff import DeepDiff

logger = logging.getLogger(__name__)


class AcademicYearViewSet(viewsets.ModelViewSet):
    queryset = AcademicYear.objects.all()
    serializer_class = AcademicYearSerializer
    permission_classes = [HasPermission]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset().order_by('-start_date')  # Order by start_date instead of year
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'status': 'success',
            'data': serializer.data
        })

    @action(detail=False, methods=['GET'])
    def current(self, request):
        """Get the current academic year"""
        try:
            current_year = AcademicYear.objects.get(is_current=True)
            serializer = self.get_serializer(current_year)
            return Response({
                'status': 'success',
                'data': serializer.data
            })
        except AcademicYear.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'No current academic year set'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'])
    def set_current(self, request, pk=None):
        academic_year = self.get_object()
        academic_year.is_current = True
        academic_year.save()
        return Response({'status': 'academic year set as current'})

class TemplateViewSet(viewsets.ModelViewSet):
    queryset = Template.objects.all()
    serializer_class = TemplateSerializer
    permission_classes = [HasPermission]
    lookup_field='code'
    
    def get_permissions(self):
        """
        Instantiates and returns the list of permissions that this view requires.
        """
        if self.action in ['approve_submission', 'reject_submission']:
            permission_classes = [HasPermission, IsIQACDirector]
        else:
            permission_classes = [HasPermission]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        queryset = Template.objects.all()
        
        # Get query parameters
        board_code = self.request.query_params.get('board')  # We're receiving board code
        academic_year = self.request.query_params.get('academic_year')
        
        board = Board.objects.filter(id=board_code).first()
        ac = AcademicYear.objects.filter(id=academic_year).first()
        print(f"Board: {board}")
        print(f"Academic Year: {ac}")


        if board_code:
            try:
                # Filter by board code instead of id
                queryset = queryset.filter(criteria__board__name=board.name)
                print(f"Templates after board filter: {queryset.count()}")
            except Exception as e:
                print(f"Error filtering by board: {str(e)}")
                raise

        # Filter by academic year if needed
        # if academic_year:
        #     try:
        #         queryset = queryset.filter(
        #             datasubmission__academic_year=ac
        #         ).distinct()
        #         print(f"Templates after academic year filter: {queryset.count()}")
        #     except Exception as e:
        #         print(f"Error filtering by academic year: {str(e)}")
        #         raise

        # print(f"Final query: {queryset.query}")
        return queryset.order_by('code')

    def list(self, request, *args, **kwargs):
        try:
            queryset = self.get_queryset()
            print(f"Total templates found: {queryset.count()}")
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        except Exception as e:
            print(f"Error in list method: {str(e)}")
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def create(self, request, *args, **kwargs):
        try:
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def get_object(self):
        try:
            # Get the lookup value from kwargs
            lookup_value = self.kwargs.get(self.lookup_field)
            # print(f"Attempting to find template with code: {lookup_value}")
            
            # Get all templates (for debugging)
            all_templates = list(self.queryset.values_list('code', flat=True))
            # print(f"Available template codes: {all_templates}")
            
            # Try to get the object
            obj = self.queryset.get(code=lookup_value)
            # print(f"Found template: {obj.code}")
            
            # Check permissions
            self.check_object_permissions(self.request, obj)
            return obj
            
        except Template.DoesNotExist:
            print(f"Template not found with code: {lookup_value}")
            raise NotFound(detail=f"Template with code '{lookup_value}' not found")

    def retrieve(self, request, *args, **kwargs):
        try:
            print("Retrieve method called")
            # instance = self.get_object()
            # print(f"Retrieved template: {instance.code}")

            # Get the filtered queryset
            queryset = self.get_queryset()
            
            # Get the lookup value (code)
            lookup_value = self.kwargs.get(self.lookup_field)
            
            # Get single object from the filtered queryset
            instance = queryset.get(code=lookup_value)

            serializer = self.get_serializer(instance)
            # print length
            print(f"Length of metadata: {len(serializer.data['metadata'])}")
            return Response(serializer.data)
        except NotFound as e:
            return Response(
                {"detail": str(e)},
                status=status.HTTP_404_NOT_FOUND
            )

    def update(self, request, code=None):
        """Update a template"""
        try:
            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get template and filter by board criteria
            queryset = Template.objects.all()
            if board_code:
                queryset = queryset.filter(criteria__board__name=board)

            template = queryset.get(code=code)

            # Check if there are any submissions for this template
            has_submissions = DataSubmission.objects.filter(
                template=template,
                academic_year=academic_year
            ).exists()

            if has_submissions:
                return Response({
                    'status': 'error',
                    'message': 'Cannot update template with existing submissions'
                }, status=status.HTTP_400_BAD_REQUEST)

            serializer = self.get_serializer(template, data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)

            return Response({
                'status': 'success',
                'message': 'Template updated successfully',
                'data': serializer.data
            })

        except Template.DoesNotExist:
            return Response({
                'status': 'error',
                'message': f"Template with code '{code}' not found for the specified board"
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def partial_update(self, request, code=None):
        """Partially update a template"""
        try:
            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get template and filter by board criteria
            queryset = Template.objects.all()
            if board_code:
                queryset = queryset.filter(criteria__board__name=board)

            template = queryset.get(code=code)

            # Check if there are any submissions for this template
            has_submissions = DataSubmission.objects.filter(
                template=template,
                academic_year=academic_year
            ).exists()

            if has_submissions:
                return Response({
                    'status': 'error',
                    'message': 'Cannot update template with existing submissions'
                }, status=status.HTTP_400_BAD_REQUEST)

            serializer = self.get_serializer(template, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)

            return Response({
                'status': 'success',
                'message': 'Template updated successfully',
                'data': serializer.data
            })

        except Template.DoesNotExist:
            return Response({
                'status': 'error',
                'message': f"Template with code '{code}' not found for the specified board"
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, code=None):
        """Delete a template"""
        try:
            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get template and filter by board criteria
            queryset = Template.objects.all()
            if board_code:
                queryset = queryset.filter(criteria__board__name=board)

            template = queryset.get(code=code)

            # Check if there are any submissions for this template
            has_submissions = DataSubmission.objects.filter(
                template=template,
                academic_year=academic_year
            ).exists()

            if has_submissions:
                return Response({
                    'status': 'error',
                    'message': 'Cannot delete template with existing submissions'
                }, status=status.HTTP_400_BAD_REQUEST)

            template.delete()

            return Response({
                'status': 'success',
                'message': 'Template deleted successfully'
            })

        except Template.DoesNotExist:
            return Response({
                'status': 'error',
                'message': f"Template with code '{code}' not found for the specified board"
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post', 'get'])
    def data(self, request, code=None):
        print("Data action called")
        try:
            template = self.get_object()
            current_year = AcademicYear.objects.filter(is_current=True).first()

            if not current_year:
                return Response({
                    'status': 'error',
                    'message': 'No active academic year found'
                }, status=status.HTTP_400_BAD_REQUEST)

            if not request.user.department:
                return Response({
                    'status': 'error',
                    'message': 'User has no associated department'
                }, status=status.HTTP_400_BAD_REQUEST)

            if request.method == 'POST':
                try:
                    with transaction.atomic():
                        # Create or get submission
                        submission, _ = DataSubmission.objects.get_or_create(
                            template=template,
                            department=request.user.department,
                            academic_year=current_year,
                            defaults={
                                'submitted_by': request.user,
                                'status': 'draft'
                            }
                        )

                        # Get the data from request
                        data = request.data

                        # Get all required fields from template metadata
                        required_fields = []
                        for section in template.metadata:
                            flat_columns = template._flatten_column_names(section['columns'])
                            required_fields.extend([
                                name for name, col in flat_columns.items()
                                if col.get('required', False)
                            ])

                        # Validate required fields
                        missing_fields = [
                            field for field in required_fields 
                            if field not in data or not data[field]
                        ]

                        if missing_fields:
                            return Response({
                                'status': 'error',
                                'message': f'Missing required fields: {", ".join(missing_fields)}'
                            }, status=status.HTTP_400_BAD_REQUEST)

                        # Add new data row
                        row_number = SubmissionData.objects.filter(
                            submission=submission
                        ).count() + 1

                        # Create submission data
                        submission_data = SubmissionData.objects.create(
                            submission=submission,
                            section_index=0,  # Assuming first section for now
                            row_number=row_number,
                            data=data
                        )

                        return Response({
                            'status': 'success',
                            'message': 'Data saved successfully',
                            'data': {
                                'id': submission_data.id,
                                'row_number': row_number,
                                'data': data
                            }
                        })

                except Exception as e:
                    return Response({
                        'status': 'error',
                        'message': str(e)
                    }, status=status.HTTP_400_BAD_REQUEST)

            # GET method
            try:
                submission = DataSubmission.objects.get(
                    template=template,
                    department=request.user.department,
                    academic_year=current_year
                )
                data_rows = SubmissionData.objects.filter(submission=submission)\
                    .order_by('row_number')
                
                return Response({
                    'status': 'success',
                    'data': {
                        'submission_id': submission.id,
                        'status': submission.status,
                        'rows': [
                            {
                                'id': row.id,
                                'row_number': row.row_number,
                                'data': row.data
                            } for row in data_rows
                        ]
                    }
                })
            except DataSubmission.DoesNotExist:
                return Response({
                    'status': 'success',
                    'data': {
                        'submission_id': None,
                        'status': None,
                        'rows': []
                    }
                })

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
    @action(detail=True, methods=['put', 'delete'])
    def data_row(self, request, code=None, *args, **kwargs):
        template = self.get_object()
        row_id = request.query_params.get('row_id')
        
        if not row_id:
            return Response({
                'status': 'error',
                'message': 'row_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            submission_data = SubmissionData.objects.get(
                id=row_id,
                submission__template=template,
                submission__department=request.user.department
            )
        except SubmissionData.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Data row not found'
            }, status=status.HTTP_404_NOT_FOUND)

        # Check if submission is editable
        if submission_data.submission.status not in ['draft', 'rejected']:
            return Response({
                'status': 'error',
                'message': 'Cannot modify data that has been submitted or approved'
            }, status=status.HTTP_400_BAD_REQUEST)

        if request.method == 'PUT':
            try:
                with transaction.atomic():
                    # Validate incoming data against template columns
                    data = request.data.get('data', {})
                    required_fields = [
                        col['name'] for col in template.columns 
                        if col.get('required', False)
                    ]
                    
                    missing_fields = [
                        field for field in required_fields 
                        if field not in data or not data[field]
                    ]
                    
                    if missing_fields:
                        return Response({
                            'status': 'error',
                            'message': f'Missing required fields: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)

                    # Update the data
                    submission_data.data = data
                    submission_data.save()

                    return Response({
                        'status': 'success',
                        'message': 'Data updated successfully',
                        'data': {
                            'id': submission_data.id,
                            'row_number': submission_data.row_number,
                            'data': submission_data.data
                        }
                    })

            except Exception as e:
                return Response({
                    'status': 'error',
                    'message': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)

        elif request.method == 'DELETE':
            try:
                with transaction.atomic():
                    # Store the row number before deleting
                    deleted_row_number = submission_data.row_number
                    submission_data.delete()

                    # Reorder remaining rows
                    subsequent_rows = SubmissionData.objects.filter(
                        submission=submission_data.submission,
                        row_number__gt=deleted_row_number
                    ).order_by('row_number')

                    for row in subsequent_rows:
                        row.row_number -= 1
                        row.save()

                    return Response({
                        'status': 'success',
                        'message': 'Data row deleted successfully'
                    })

            except Exception as e:
                return Response({
                    'status': 'error',
                    'message': str(e)
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            'status': 'error',
            'message': 'Invalid method'
        }, status=status.HTTP_405_METHOD_NOT_ALLOWED)
    
    @action(detail=True, methods=['get', 'post'], url_path='sections/(?P<section_index>\d+)/data')
    def section_data(self, request, code=None, section_index=None):
        """Handle section-specific data operations"""
        try:
            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            print(f"Board: {board}")
            print(f"Academic Year: {academic_year}")

            # Get template and filter by board criteria
            queryset = Template.objects.all()
            if board_code:
                try:
                    queryset = queryset.filter(criteria__board__name=board)
                    print(f"Templates after board filter: {queryset.count()}")
                except Exception as e:
                    print(f"Error filtering by board: {str(e)}")
                    raise

            try:
                template = queryset.get(code=code)
            except Template.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': f"Template with code '{code}' not found for the specified board"
                }, status=status.HTTP_404_NOT_FOUND)

            department = Department.objects.get(id=request.user.department_id)
            print('~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~')
            print(f"Department: {department}")
            section_index = int(section_index)

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            if not request.user.department:
                return Response({
                    'status': 'error',
                    'message': 'User has no associated department'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Validate section index
            if section_index >= len(template.metadata):
                return Response({
                    'status': 'error',
                    'message': 'Invalid section index'
                }, status=status.HTTP_400_BAD_REQUEST)

            if request.method == 'GET':
                try:
                    submission = DataSubmission.objects.get(
                        template=template,
                        department=department,
                        academic_year=academic_year
                    )
                    data_rows = SubmissionData.objects.filter(
                        submission=submission,
                        section_index=section_index
                    ).order_by('row_number')
                    
                    return Response({
                        'status': 'success',
                        'data': {
                            'submission_id': submission.id,
                            'status': submission.status,
                            'rows': [
                                {
                                    'id': row.id,
                                    'row_number': row.row_number,
                                    'data': row.data
                                } for row in data_rows
                            ]
                        }
                    })
                except DataSubmission.DoesNotExist:
                    return Response({
                        'status': 'success',
                        'data': {
                            'submission_id': None,
                            'status': None,
                            'rows': []
                        }
                    })

            elif request.method == 'POST':
                try:
                    with transaction.atomic():
                        # Get or create submission without creating data rows
                        submission, _ = DataSubmission.objects.get_or_create(
                            template=template,
                            department=department,
                            academic_year=academic_year,
                            defaults={
                                'submitted_by': request.user,
                                'status': 'draft'
                            }
                        )

                        # Validate incoming data against template structure
                        section = template.metadata[section_index]
                        data = request.data
                        print(f"Data: {data}")
                        
                        # Validate required fields
                        required_fields = []
                        for column in section['columns']:
                            if column.get('required', False):
                                if column['type'] == 'single':
                                    required_fields.append(column['name'])
                                elif column['type'] == 'group':
                                    for nested_col in column['columns']:
                                        if nested_col.get('required', False):
                                            required_fields.append(f"{column['name']}_{nested_col['name']}")

                        missing_fields = [
                            field for field in required_fields 
                            if field not in data or not data[field]
                        ]

                        if missing_fields:
                            return Response({
                                'status': 'error',
                                'message': f'Missing required fields: {", ".join(missing_fields)}'
                            }, status=status.HTTP_400_BAD_REQUEST)

                        # Get the next row number for this specific section
                        row_number = SubmissionData.objects.filter(
                            submission=submission,
                            section_index=section_index
                        ).count() + 1
                        
                        # Create data row only for this section
                        submission_data = SubmissionData.objects.create(
                            submission=submission,
                            section_index=section_index,
                            row_number=row_number,
                            data=data
                        )

                        return Response({
                            'status': 'success',
                            'message': 'Data saved successfully',
                            'data': {
                                'id': submission_data.id,
                                'row_number': row_number,
                                'data': data
                            }
                        })

                except Exception as e:
                    return Response({
                        'status': 'error',
                        'message': str(e)
                    }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    @action(detail=True, methods=['put', 'delete'], url_path=r'sections/(?P<section_index>\d+)/data/(?P<row_id>\d+)')
    def section_data_row(self, request, code=None, section_index=None, row_id=None):
        """
        Handle section-specific row operations
        
        PUT: Update a specific row in a section
        DELETE: Delete a specific row from a section
        """
        # Convert section_index and row_id to int
        section_index = int(section_index)
        row_id = int(row_id)
        
        try:
            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            print(f"Board: {board}")
            print(f"Academic Year: {academic_year}")

            # Get template and filter by board criteria
            queryset = Template.objects.all()
            if board_code:
                try:
                    queryset = queryset.filter(criteria__board__name=board)
                    print(f"Templates after board filter: {queryset.count()}")
                except Exception as e:
                    print(f"Error filtering by board: {str(e)}")
                    raise

            try:
                template = queryset.get(code=code)
            except Template.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': f"Template with code '{code}' not found for the specified board"
                }, status=status.HTTP_404_NOT_FOUND)

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)
                
            try:
                submission_data = SubmissionData.objects.get(
                    id=row_id,
                    submission__template=template,
                    submission__department=request.user.department,
                    submission__academic_year=academic_year,
                    section_index=section_index
                )
            except SubmissionData.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': 'Data row not found'
                }, status=status.HTTP_404_NOT_FOUND)

            # Check if submission is editable
            if submission_data.submission.status not in ['draft', 'rejected']:
                return Response({
                    'status': 'error',
                    'message': 'Cannot modify data that has been submitted or approved'
                }, status=status.HTTP_400_BAD_REQUEST)

            if request.method == 'PUT':
                try:
                    with transaction.atomic():
                        # Validate incoming data
                        section = template.metadata[section_index]
                        data = request.data.get('data', {})
                        
                        # Validate required fields (same validation as POST)
                        required_fields = []
                        for column in section['columns']:
                            if column.get('required', False):
                                if column['type'] == 'single':
                                    required_fields.append(column['name'])
                                elif column['type'] == 'group':
                                    for nested_col in column['columns']:
                                        if nested_col.get('required', False):
                                            required_fields.append(f"{column['name']}_{nested_col['name']}")

                        missing_fields = [
                            field for field in required_fields 
                            if field not in data or not data[field]
                        ]

                        if missing_fields:
                            return Response({
                                'status': 'error',
                                'message': f'Missing required fields: {", ".join(missing_fields)}'
                            }, status=status.HTTP_400_BAD_REQUEST)

                        # Update the data
                        submission_data.data = data
                        submission_data.save()

                        return Response({
                            'status': 'success',
                            'message': 'Data updated successfully',
                            'data': {
                                'id': submission_data.id,
                                'row_number': submission_data.row_number,
                                'data': submission_data.data
                            }
                        })

                except Exception as e:
                    return Response({
                        'status': 'error',
                        'message': str(e)
                    }, status=status.HTTP_400_BAD_REQUEST)

            elif request.method == 'DELETE':
                try:
                    with transaction.atomic():
                        # Store the row number before deleting
                        deleted_row_number = submission_data.row_number
                        submission_data.delete()

                        # Reorder remaining rows
                        subsequent_rows = SubmissionData.objects.filter(
                            submission=submission_data.submission,
                            section_index=section_index,
                            row_number__gt=deleted_row_number
                        ).order_by('row_number')

                        for row in subsequent_rows:
                            row.row_number -= 1
                            row.save()

                        return Response({
                            'status': 'success',
                            'message': 'Data row deleted successfully'
                        })

                except Exception as e:
                    return Response({
                        'status': 'error',
                        'message': str(e)
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    @action(detail=True, methods=['get'], url_path='submission')
    def submission_state(self, request, code=None):
        """Get submission state for template based on board and academic year"""
        try:
            # Get query parameters
            board_id = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            if not board_id or not academic_year_id:
                return Response({
                    'status': 'error',
                    'message': 'Board ID and Academic Year ID are required'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get template for the specific board
            template = Template.objects.filter(
                code=code,
                criteria__board_id=board_id
            ).first()

            if not template:
                return Response({
                    'status': 'error',
                    'message': f"Template with code '{code}' not found for the specified board"
                }, status=status.HTTP_404_NOT_FOUND)

            # Get department from user
            department = getattr(request.user, 'department', None)
            if not department:
                return Response({
                    'status': 'error',
                    'message': 'User must be associated with a department'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get submission if it exists
            submission = DataSubmission.objects.filter(
                template=template,
                academic_year_id=academic_year_id,
                department=department
            ).first()

            if submission:
                return Response({
                    'status': 'success',
                    'data': {
                        'id': submission.id,
                        'status': submission.status,
                        'submitted_at': submission.submitted_at,
                        'rejection_reason': submission.rejection_reason,
                        'can_edit': submission.can_be_edited()
                    }
                })
            else:
                return Response({
                    'status': 'success',
                    'data': {
                        'id': None,
                        'status': 'draft',
                        'submitted_at': None,
                        'rejection_reason': None,
                        'can_edit': True
                    }
                })

        except Exception as e:
            print(f"Error in submission_state: {str(e)}")
            return Response({
                'status': 'error',
                'message': 'An unexpected error occurred'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _process_column_row(self, row):
        """Process a row of column definitions"""
        columns = []
        for cell in row:
            if cell.value:
                column = {
                    'name': cell.value.strip(),
                    'type': 'single',
                    'data_type': self._determine_data_type(cell.value),
                    'required': True
                }
                
                if column['data_type'] == 'option':
                    column['options'] = self._determine_options(cell.value)
                
                columns.append(column)
        return columns

    def _determine_data_type(self, value):
        """Determine the data type of a column based on its name"""
        value_lower = value.lower()
        if 'date' in value_lower:
            return 'date'
        elif 'email' in value_lower:
            return 'email'
        elif 'link' in value_lower or 'url' in value_lower:
            return 'url'
        elif 'number' in value_lower or 'amount' in value_lower:
            return 'number'
        elif '(yes/no)' in value_lower or any(opt in value_lower for opt in ['choose', 'select']):
            return 'option'
        return 'string'

    def _determine_options(self, value):
        """Determine options for option-type columns"""
        if '(yes/no)' in value.lower():
            return ['Yes', 'No']
        # Add more option patterns as needed
        return []

    def _is_group_header(self, columns):
        """Determine if a row of columns represents a group header"""
        return len(columns) == 1 and not any(
            keyword in columns[0]['name'].lower() 
            for keyword in ['link', 'url', 'email', 'date']
        )

    def _create_group_column(self, columns):
        """Create a group column structure"""
        return {
            'name': columns[0]['name'],
            'type': 'group',
            'columns': []
        }

    def _process_column_groups(self, row):
        """Process row containing column group headers"""
        column_groups = []
        current_group = None
        
        for cell in row:
            if cell.value:
                current_group = cell.value.strip()
            if current_group:
                column_groups.append(current_group)
            
        return column_groups

    def _create_columns_with_groups(self, row, column_groups):
        """Create column definitions with their groups"""
        columns = []
        for idx, cell in enumerate(row):
            if cell.value:
                column_name = cell.value.strip()
                group = column_groups[idx] if idx < len(column_groups) else None
                
                # Create column definition
                column = {
                    'name': f"{column_name.lower().replace(' ', '_')}",
                    'display_name': column_name,
                    'type': 'number',  # Default to number for this template
                    'required': True,
                    'description': f'Enter {column_name.lower()}',
                    'group': group
                }
                
                columns.append(column)
        
        return columns
    
    # @action(detail=True, methods=['get', 'post'], url_path='submission')
    # def submission_state(self, request, code=None):
    #     """Get or create submission state for template"""
    #     template = self.get_object()
    #     current_year = AcademicYear.objects.filter(is_current=True).first()
    #     print()
    #     # Attempt to retrieve the Department instance
    #     try:
    #         department_name = request.user.department
    #         department = Department.objects.get(id=request.user.department_id)
    #     except Department.DoesNotExist:
    #         return Response({
    #             'status': 'error',
    #             'message': f'Department "{request.user.department}" not found.'
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     # Check for current academic year
    #     if not current_year:
    #         return Response({
    #             'status': 'error',
    #             'message': 'No active academic year found'
    #         }, status=status.HTTP_400_BAD_REQUEST)

    #     try:
    #         # Check if a submission already exists
    #         submission = DataSubmission.objects.get(
    #             template=template,
    #             department=department,
    #             academic_year=current_year,
    #             # board=board
    #         )
    #         serializer = DataSubmissionSerializer(submission)
    #         return Response({
    #             'status': 'success',
    #             'data': serializer.data
    #         })
    #     except DataSubmission.DoesNotExist:
    #         # Create a new submission if the request is POST
    #         if request.method == 'POST':
    #             submission = DataSubmission.objects.create(
    #                 template=template,
    #                 department=department,
    #                 academic_year=current_year,
    #                 submitted_by=request.user,
    #                 # board=Board.objects.get(id=request.data.get('board')),
    #                 status='draft'
    #             )
    #             serializer = DataSubmissionSerializer(submission)
    #             return Response({
    #                 'status': 'success',
    #                 'data': serializer.data
    #             })

    #         # Return None if no submission exists and the request is GET
    #         return Response({
    #             'status': 'success',
    #             'data': None
    #         })
    #         try:
    #             submission = DataSubmission.objects.get(
    #                 template=template,
    #                 department=request.user.department,
    #                 academic_year=academic_year,
    #             )
    #             serializer = DataSubmissionSerializer(submission)
    #             return Response({
    #                 'status': 'success',
    #                 'data': serializer.data
    #             })
    #         except DataSubmission.DoesNotExist:
    #             if request.method == 'POST':
    #                 submission = DataSubmission.objects.create(
    #                     template=template,
    #                     department=request.user.department,
    #                     academic_year=academic_year,
    #                     submitted_by=request.user,
    #                     status='draft'
    #                 )
    #                 serializer = DataSubmissionSerializer(submission)
    #                 return Response({
    #                     'status': 'success',
    #                     'data': serializer.data
    #                 })
    #             return Response({
    #                 'status': 'success',
    #                 'data': None
    #             })

    #     except Exception as e:
    #         return Response({
    #             'status': 'error',
    #             'message': str(e)
    #         }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='submit')
    def submit_template(self, request, code=None):
        """Submit template for review"""
        try:
            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            print(f"Board: {board}")
            print(f"Academic Year: {academic_year}")

            # Get template and filter by board criteria
            queryset = Template.objects.all()
            if board_code:
                try:
                    queryset = queryset.filter(criteria__board__name=board)
                    print(f"Templates after board filter: {queryset.count()}")
                except Exception as e:
                    print(f"Error filtering by board: {str(e)}")
                    raise

            try:
                template = queryset.get(code=code)
            except Template.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': f"Template with code '{code}' not found for the specified board"
                }, status=status.HTTP_404_NOT_FOUND)

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                submission = DataSubmission.objects.get(
                    template=template,
                    department=request.user.department,
                    academic_year=academic_year
                )
                
                if submission.status not in ['draft', 'rejected']:
                    return Response({
                        'status': 'error',
                        'message': 'Only draft or rejected submissions can be submitted'
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Validate that all required sections have data
                for section_index, section in enumerate(template.metadata):
                    data_rows = SubmissionData.objects.filter(
                        submission=submission,
                        section_index=section_index
                    ).count()
                    
                    if section.get('required', False) and data_rows == 0:
                        return Response({
                            'status': 'error',
                            'message': f'Required section "{section.get("title", f"Section {section_index + 1}")}" has no data'
                        }, status=status.HTTP_400_BAD_REQUEST)

                submission.status = 'submitted'
                submission.submitted_at = timezone.now()
                submission.save()

                return Response({
                    'status': 'success',
                    'message': 'Template submitted successfully',
                    'data': DataSubmissionSerializer(submission).data
                })
            except DataSubmission.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': f'No submission found for academic year {academic_year}'
                }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    @action(detail=True, methods=['post'], url_path='withdraw')
    def withdraw_submission(self, request, code=None):
        """Withdraw submitted template"""
        try:
            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            # Get template and filter by board criteria
            queryset = Template.objects.all()
            if board_code:
                queryset = queryset.filter(criteria__board__name=board)

            template = queryset.get(code=code)

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                submission = DataSubmission.objects.get(
                    template=template,
                    department=request.user.department,
                    academic_year=academic_year
                )
                
                if submission.status != 'submitted':
                    return Response({
                        'status': 'error',
                        'message': 'Only submitted templates can be withdrawn'
                    }, status=status.HTTP_400_BAD_REQUEST)

                submission.status = 'draft'
                submission.save()

                return Response({
                    'status': 'success',
                    'message': 'Submission withdrawn successfully',
                    'data': DataSubmissionSerializer(submission).data
                })
            except DataSubmission.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': f'No submission found for academic year {academic_year}'
                }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve_submission(self, request, code=None):
        """Approve a submission (IQAC Director only)"""
        if not request.user.roles.filter(name='IQAC Director').exists():
            return Response({
                'status': 'error',
                'message': 'Only IQAC Director can approve submissions'
            }, status=status.HTTP_403_FORBIDDEN)

        try:
            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            # Get template and filter by board criteria
            queryset = Template.objects.all()
            if board_code:
                queryset = queryset.filter(criteria__board__name=board)

            template = queryset.get(code=code)

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            department = request.data.get('department')
            if not department:
                return Response({
                    'status': 'error',
                    'message': 'Department is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                submission = DataSubmission.objects.get(
                    template=template,
                    department=department,
                    academic_year=academic_year
                )
                
                if submission.status != 'submitted':
                    return Response({
                        'status': 'error',
                        'message': 'Only submitted data can be approved'
                    }, status=status.HTTP_400_BAD_REQUEST)

                submission.status = 'approved'
                submission.verified_by = request.user
                submission.verified_at = timezone.now()
                submission.save()

                return Response({
                    'status': 'success',
                    'message': 'Submission approved successfully',
                    'data': DataSubmissionSerializer(submission).data
                })
            except DataSubmission.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': f'No submission found for department {department} and academic year {academic_year}'
                }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject_submission(self, request, code=None):
        """Reject a submission with comments (IQAC Director only)"""
        if not request.user.roles.filter(name='IQAC Director').exists():
            return Response({
                'status': 'error',
                'message': 'Only IQAC Director can reject submissions'
            }, status=status.HTTP_403_FORBIDDEN)

        try:
            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            # Get template and filter by board criteria
            queryset = Template.objects.all()
            if board_code:
                queryset = queryset.filter(criteria__board__name=board)

            template = queryset.get(code=code)

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            department = request.data.get('department')
            if not department:
                return Response({
                    'status': 'error',
                    'message': 'Department is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                submission = DataSubmission.objects.get(
                    template=template,
                    department=department,
                    academic_year=academic_year
                )
                
                if submission.status != 'submitted':
                    return Response({
                        'status': 'error',
                        'message': 'Only submitted data can be rejected'
                    }, status=status.HTTP_400_BAD_REQUEST)

                reason = request.data.get('reason')
                if not reason:
                    return Response({
                        'status': 'error',
                        'message': 'Rejection reason is required'
                    }, status=status.HTTP_400_BAD_REQUEST)

                submission.status = 'rejected'
                submission.verified_by = request.user
                submission.verified_at = timezone.now()
                submission.rejection_reason = reason
                submission.save()

                return Response({
                    'status': 'success',
                    'message': 'Submission rejected successfully',
                    'data': DataSubmissionSerializer(submission).data
                })
            except DataSubmission.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': f'No submission found for department {department} and academic year {academic_year}'
                }, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    @action(detail=False, methods=['POST'])
    def import_from_excel(self, request):
        """Import template from Excel file with board and academic year filtering"""
        try:
            # Validate file
            if 'file' not in request.FILES:
                return Response({
                    'status': 'error',
                    'message': 'No file provided'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get query parameters
            board_code = request.query_params.get('board')
            academic_year_id = request.query_params.get('academic_year')

            # Get board and academic year instances
            board = Board.objects.filter(id=board_code).first()
            academic_year = AcademicYear.objects.filter(id=academic_year_id).first()

            if not board:
                return Response({
                    'status': 'error',
                    'message': 'Board not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'Academic year not found'
                }, status=status.HTTP_400_BAD_REQUEST)

            file = request.FILES['file']
            template_code = file.name.split('.xlsx')[0]

            # Check if template exists for this board and has submissions
            try:
                existing_template = Template.objects.get(
                    code=template_code,
                    criteria__board=board
                )
                
                has_submissions = DataSubmission.objects.filter(
                    template=existing_template,
                    academic_year=academic_year
                ).exists()

                if has_submissions:
                    return Response({
                        'status': 'error',
                        'message': 'Cannot update template with existing submissions'
                    }, status=status.HTTP_400_BAD_REQUEST)
            except Template.DoesNotExist:
                pass

            with transaction.atomic():
                wb = openpyxl.load_workbook(file)
                ws = wb.active

                # Process Excel structure
                sections = []
                current_section = None
                current_headers = []
                current_group = None
                column_groups = []
                
                for row_index, row in enumerate(ws.rows, start=1):
                    first_cell = row[0].value
                    if not first_cell:
                        continue

                    # Check if this is a header row
                    if re.match(r'^\d+\.\d+\.?\d*\s+[a-zA-Z]', str(first_cell)):
                        if current_section:
                            sections.append(current_section)
                        
                        current_headers = [first_cell.strip()]
                        current_section = {
                            'headers': current_headers,
                            'columns': []
                        }
                    
                    # Check if this is a column group row
                    elif first_cell and any(cell.value for cell in row[1:]):
                        # Check if this row contains group headers
                        if self._is_group_row(row):
                            column_groups = self._process_column_groups(row)
                        else:
                            columns = self._process_column_row(row)
                            if column_groups:
                                columns = self._create_columns_with_groups(row, column_groups)
                            
                            if self._is_group_header(columns):
                                current_group = self._create_group_column(columns)
                                current_section['columns'].append(current_group)
                            else:
                                if current_group:
                                    current_group['columns'].extend(columns)
                                    current_group = None
                                else:
                                    current_section['columns'].extend(columns)

                # Add the last section
                if current_section:
                    sections.append(current_section)

                # Get or create criteria for the board
                criteria, _ = Criteria.objects.get_or_create(
                    board=board,
                    defaults={'name': f'Default Criteria for {board.name}'}
                )

                # Create template data
                template_data = {
                    'code': template_code,
                    'name': sections[0]['headers'][0] if sections else '',
                    'metadata': sections,
                    'criteria': criteria.id
                }

                # Save or update template
                try:
                    template = Template.objects.get(
                        code=template_code,
                        criteria__board=board
                    )
                    serializer = self.get_serializer(template, data=template_data)
                except Template.DoesNotExist:
                    serializer = self.get_serializer(data=template_data)

                if serializer.is_valid():
                    template = serializer.save()
                    
                    # Create or update template version
                    TemplateVersion.objects.create(
                        template=template,
                        academic_year=academic_year,
                        version_data=template_data
                    )

                    return Response({
                        'status': 'success',
                        'message': f'Successfully imported template {template_code} for {board.name}',
                        'data': serializer.data
                    })
                else:
                    return Response({
                        'status': 'error',
                        'message': 'Invalid template data',
                        'errors': serializer.errors
                    }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({
                'status': 'error',
                'message': f'Error processing template: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _process_column_row(self, row):
        """Process a row of column definitions"""
        columns = []
        for cell in row:
            if cell.value:
                column = {
                    'name': cell.value.strip(),
                    'type': 'single',
                    'data_type': self._determine_data_type(cell.value),
                    'required': True
                }
                
                if column['data_type'] == 'option':
                    column['options'] = self._determine_options(cell.value)
                
                columns.append(column)
        return columns

    def _determine_data_type(self, value):
        """Determine the data type of a column based on its name"""
        value_lower = str(value).lower()
        if 'date' in value_lower:
            return 'date'
        elif 'email' in value_lower:
            return 'email'
        elif 'link' in value_lower or 'url' in value_lower:
            return 'url'
        elif any(term in value_lower for term in ['number', 'amount', 'count', 'total', 'percentage']):
            return 'number'
        elif '(yes/no)' in value_lower or any(opt in value_lower for opt in ['choose', 'select']):
            return 'option'
        return 'string'

    def _determine_options(self, value):
        """Determine options for option-type columns"""
        value_lower = str(value).lower()
        if '(yes/no)' in value_lower:
            return ['Yes', 'No']
        elif 'priority' in value_lower:
            return ['High', 'Medium', 'Low']
        elif 'status' in value_lower:
            return ['Active', 'Inactive', 'Pending']
        return []

    def _is_group_row(self, row):
        """Determine if a row contains group headers"""
        non_empty_cells = [cell.value for cell in row if cell.value]
        if len(non_empty_cells) <= 1:
            return False
        # Check if any cell contains detailed data
        return not any(self._has_data_indicators(str(cell)) for cell in non_empty_cells)

    def _has_data_indicators(self, value):
        """Check if a value contains indicators of detailed data"""
        indicators = ['date', 'email', 'number', 'amount', '(yes/no)', 'link', 'url']
        return any(indicator in value.lower() for indicator in indicators)

    def _is_group_header(self, columns):
        """Determine if a row of columns represents a group header"""
        return len(columns) == 1 and not any(
            keyword in columns[0]['name'].lower() 
            for keyword in ['link', 'url', 'email', 'date']
        )

    def _create_group_column(self, columns):
        """Create a group column structure"""
        return {
            'name': columns[0]['name'],
            'type': 'group',
            'columns': []
        }

    def _process_column_groups(self, row):
        """Process row containing column group headers"""
        column_groups = []
        current_group = None
        
        for cell in row:
            if cell.value:
                current_group = cell.value.strip()
            if current_group:
                column_groups.append(current_group)
        
        return column_groups

    def _create_columns_with_groups(self, row, column_groups):
        """Create column definitions with their groups"""
        columns = []
        for idx, cell in enumerate(row):
            if cell.value:
                column_name = cell.value.strip()
                group = column_groups[idx] if idx < len(column_groups) else None
                
                column = {
                    'name': self._sanitize_column_name(column_name),
                    'display_name': column_name,
                    'type': self._determine_data_type(column_name),
                    'required': True,
                    'description': f'Enter {column_name.lower()}',
                    'group': group
                }
                
                if column['type'] == 'option':
                    column['options'] = self._determine_options(column_name)
                
                columns.append(column)
        
        return columns

    def _sanitize_column_name(self, name):
        """Convert column name to a valid field name"""
        sanitized = re.sub(r'[^\w\s-]', '', name.lower())
        return re.sub(r'[-\s]+', '_', sanitized)
class NameAutocompleteView(APIView):
    permission_classes = [HasPermission]

    def get(self, request, *args, **kwargs):
        query = request.GET.get('q', '')
        if query:
            # Fetch distinct names that contain the query string (case-insensitive)
            # matching_names = (
            #     Template.objects.filter(name__icontains=query)
            #     .values_list('name', flat=True)
            #     .distinct()[:10]  # Limit to top 10 suggestions
            # )
            matching_names = ['John Doe', 'Jane Doe', 'Alice Smith']
            

            return Response(list(matching_names), status=status.HTTP_200_OK)
        
        return Response([], status=status.HTTP_200_OK)
    
class DataSubmissionViewSet(viewsets.ModelViewSet):
    serializer_class = DataSubmissionSerializer
    permission_classes = [HasPermission]
    filterset_class = DataSubmissionFilter
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    ordering_fields = [
        'created_at', 
        'updated_at', 
        'submitted_at',
        'verified_at',
        'academic_year__start_date',
        'department__name',
        'template__name',
        'status'
    ]
    ordering = ['-academic_year__start_date', '-updated_at']
    queryset = DataSubmission.objects.all()

    def get_queryset(self):
        user = self.request.user

        queryset = DataSubmission.objects.select_related(
            'template',
            'department',
            'academic_year',
            'submitted_by',
            'verified_by'
        ).prefetch_related('data_rows')
        # Filter based on user role
        if user.roles.filter(name='faculty').exists():
            # If the user is a faculty member, filter by department
            queryset = queryset.filter(department=user.department)
        elif user.roles.filter(name='iqac_director').exists():
            # print(user.roles.get())
            # print(user.roles.get().name)
            # IQAC Director can see all submissions (no additional filter applied)
            pass
        else:
            # For other roles, return an empty queryset
            queryset = queryset.none()

        return queryset

    # Add this method to customize filter queryset
    def filter_queryset(self, queryset):
        queryset = super().filter_queryset(queryset)
        
        # board = self.request.query_params.get('board')
        # if board:
        #     queryset = queryset.filter(board_id=board)
        # Additional custom filtering
        if self.request.query_params.get('is_current_year'):
            queryset = queryset.filter(academic_year__is_current=True)
            
        search_query = self.request.query_params.get('search')
        if search_query:
            queryset = queryset.filter(
                models.Q(template__name__icontains=search_query) |
                models.Q(department__name__icontains=search_query) |
                models.Q(submitted_by__username__icontains=search_query) |
                models.Q(template__code__icontains=search_query)
            )
            
        return queryset

    def perform_create(self, serializer):
        submission = serializer.save(submitted_by=self.request.user)
        SubmissionHistory.objects.create(
            submission=submission,
            action='created',
            performed_by=self.request.user
        )
        
    @action(detail=False, methods=['get'], url_path='current_academic_year')
    def current_academic_year(self, request):
        """Get submissions for current academic year"""
        
        board = Board.objects.filter(id=request.query_params.get('board')).first()

        if not board:
            return Response({
                'status': 'error',
                'message': 'Board ID is required'
            }, status=status.HTTP_400_BAD_REQUEST)
            
        
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if not current_year:
            return Response({
                'status': 'error',
                'message': 'No current academic year set'
            }, status=status.HTTP_404_NOT_FOUND)
        print(current_year)
        queryset = self.filter_queryset(self.get_queryset().filter(
            academic_year=current_year,
            template__criteria__board=board
        ))
        # queryset = self.get_queryset()
        # print(current_year)
        page = self.paginate_queryset(queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        print('submissions: ', serializer.data)
        return Response(serializer.data)

    @action(detail=False)
    def submission_status(self, request):
        """Get submission status summary for current academic year"""
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if not current_year:
            return Response({
                'status': 'error',
                'message': 'No current academic year set'
            }, status=status.HTTP_404_NOT_FOUND)

        queryset = self.get_queryset().filter(academic_year=current_year)
        
        summary = {
            'total': queryset.count(),
            'draft': queryset.filter(status='draft').count(),
            'submitted': queryset.filter(status='submitted').count(),
            'approved': queryset.filter(status='approved').count(),
            'rejected': queryset.filter(status='rejected').count(),
            'academic_year': AcademicYearSerializer(current_year).data
        }

        return Response(summary)

    @action(detail=True, methods=['post'])
    def add_row(self, request, pk=None):
        """Add a new row to a specific section"""
        submission = self.get_object()
        
        # Check if submission is editable
        if submission.status not in ['draft', 'rejected']:
            return Response({
                'status': 'error',
                'message': 'Cannot modify data in current status'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Validate section index
        section_index = request.data.get('section_index')
        if section_index is None:
            return Response({
                'status': 'error',
                'message': 'section_index is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # Get the next row number for this section
                next_row = (SubmissionData.objects.filter(
                    submission=submission,
                    section_index=section_index
                ).count() + 1)

                # Create new submission data
                submission_data = SubmissionData.objects.create(
                    submission=submission,
                    section_index=section_index,
                    row_number=next_row,
                    data=request.data.get('data', {})
                )

                return Response({
                    'status': 'success',
                    'message': 'Row added successfully',
                    'data': SubmissionDataSerializer(submission_data).data
                })

        except ValidationError as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    def _track_changes(self, submission, previous_state, new_state, action, user, additional_details=None):
        """Track changes in submission history"""
        diff = DeepDiff(previous_state, new_state, ignore_order=True)
        
        details = {
            'changes': [],
            'additional_info': additional_details
        }

        # Process changes
        if 'values_changed' in diff:
            for path, change in diff['values_changed'].items():
                details['changes'].append({
                    'type': 'changed',
                    'path': path,
                    'old_value': change['old_value'],
                    'new_value': change['new_value']
                })

        if 'dictionary_item_added' in diff:
            for path in diff['dictionary_item_added']:
                details['changes'].append({
                    'type': 'added',
                    'path': path
                })

        if 'dictionary_item_removed' in diff:
            for path in diff['dictionary_item_removed']:
                details['changes'].append({
                    'type': 'removed',
                    'path': path
                })

        SubmissionHistory.objects.create(
            submission=submission,
            action=action,
            performed_by=user,
            details=details,
            previous_data=previous_state,
            new_data=new_state
        )

    @action(detail=True, methods=['PUT'])
    def update_row(self, request, pk=None):
        submission = self.get_object()
        
        if submission.status not in ['draft', 'rejected']:
            return Response({
                'status': 'error',
                'message': 'Cannot modify data in current status'
            }, status=status.HTTP_400_BAD_REQUEST)

        previous_state = self._get_submission_data_state(submission)

        try:
            with transaction.atomic():
                # Update the data
                row_id = request.data.get('row_id')
                new_data = request.data.get('data', {})
                
                submission_data = get_object_or_404(
                    SubmissionData, 
                    submission=submission,
                    id=row_id
                )
                
                submission_data.data = new_data
                submission_data.full_clean()
                submission_data.save()

                # Get new state and track changes
                new_state = self._get_submission_data_state(submission)
                self._track_changes(
                    submission=submission,
                    previous_state=previous_state,
                    new_state=new_state,
                    action='updated',
                    user=request.user
                )

                return Response({
                    'status': 'success',
                    'message': 'Row updated successfully',
                    'data': SubmissionDataSerializer(submission_data).data
                })

        except ValidationError as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['delete'])
    def delete_row(self, request, pk=None):
        """Delete a row and reorder remaining rows"""
        submission = self.get_object()
        
        if submission.status not in ['draft', 'rejected']:
            return Response({
                'status': 'error',
                'message': 'Cannot modify data in current status'
            }, status=status.HTTP_400_BAD_REQUEST)

        row_id = request.query_params.get('row_id')
        if not row_id:
            return Response({
                'status': 'error',
                'message': 'row_id is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                submission_data = get_object_or_404(
                    SubmissionData, 
                    submission=submission,
                    id=row_id
                )
                
                section_index = submission_data.section_index
                deleted_row_number = submission_data.row_number
                
                # Delete the row
                submission_data.delete()
                
                # Reorder remaining rows
                SubmissionData.objects.filter(
                    submission=submission,
                    section_index=section_index,
                    row_number__gt=deleted_row_number
                ).update(row_number=models.F('row_number') - 1)

                return Response({
                    'status': 'success',
                    'message': 'Row deleted successfully'
                })

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit the data for approval"""
        submission = self.get_object()
        
        if submission.status not in ['draft', 'rejected']:
            return Response({
                'status': 'error',
                'message': 'Only draft or rejected submissions can be submitted'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Validate all data rows
            for data_row in submission.data_rows.all():
                data_row.full_clean()

            submission.status = 'submitted'
            submission.submitted_at = timezone.now()
            submission.save()

            return Response({
                'status': 'success',
                'message': 'Submission successful'
            })

        except ValidationError as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['POST'])
    def approve(self, request, pk=None):
        submission = self.get_object()
        if submission.status != 'submitted':
            return Response({
                'status': 'error',
                'message': 'Only submitted data can be approved'
            }, status=status.HTTP_400_BAD_REQUEST)

        submission.status = 'approved'
        submission.verified_by = request.user
        submission.verified_at = timezone.now()
        submission.save()

        # Record history
        SubmissionHistory.objects.create(
            submission=submission,
            action='approved',
            performed_by=request.user
        )

        return Response({
            'status': 'success',
            'message': 'Submission approved'
        })

    @action(detail=True, methods=['POST'])
    def reject(self, request, pk=None):
        submission = self.get_object()
        reason = request.data.get('reason')
        
        if not reason:
            return Response({
                'status': 'error',
                'message': 'Rejection reason is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        submission.status = 'rejected'
        submission.verified_by = request.user
        submission.verified_at = timezone.now()
        submission.rejection_reason = reason
        submission.save()

        # Record history
        SubmissionHistory.objects.create(
            submission=submission,
            action='rejected',
            performed_by=request.user,
            details={'reason': reason}
        )

        return Response({
            'status': 'success',
            'message': 'Submission rejected'
        })
        
    @action(detail=False, methods=['GET'])
    def stats(self, request):
        """Get detailed submission statistics"""
        current_year = AcademicYear.objects.filter(is_current=True).first()
        board_code = request.query_params.get('board')
        
        if not current_year:
            return Response({
                'status': 'error',
                'message': 'No current academic year set'
            }, status=status.HTTP_404_NOT_FOUND)

        # Base queryset for current academic year
        base_queryset = self.get_queryset().filter(academic_year=current_year)  # Added academic year filter
        if board_code:
            base_queryset = base_queryset.filter(
                template__criteria__board__code=board_code
            )

        # Get today's date for daily stats
        today = timezone.now().date()
        yesterday = today - timezone.timedelta(days=1)
        last_week = today - timezone.timedelta(days=7)
        last_month = today - timezone.timedelta(days=30)

        # Calculate review time for approved submissions
        approved_submissions = base_queryset.filter(
            status='approved',
            verified_at__isnull=False,
            submitted_at__isnull=False  # Added this check
        )
        
        avg_review_time = approved_submissions.annotate(
            review_time=models.ExpressionWrapper(
                models.F('verified_at') - models.F('submitted_at'),
                output_field=models.DurationField()
            )
        ).aggregate(avg=models.Avg('review_time'))['avg']

        total_submissions = base_queryset.count()
        approved_count = base_queryset.filter(status='approved').count()

        stats = {
            # Status counts
            'pending': base_queryset.filter(status='submitted').count(),
            'approved': approved_count,
            'rejected': base_queryset.filter(status='rejected').count(),
            'draft': base_queryset.filter(status='draft').count(),  # Added draft count
            'total': total_submissions,
            
            # Time-based stats
            'today': base_queryset.filter(created_at__date=today).count(),
            'yesterday': base_queryset.filter(created_at__date=yesterday).count(),
            'this_week': base_queryset.filter(created_at__date__gte=last_week).count(),
            'this_month': base_queryset.filter(created_at__date__gte=last_month).count(),
            
            # Department stats
            'departments': {
                'total': base_queryset.values('department').distinct().count(),
                'active_this_month': base_queryset.filter(
                    created_at__date__gte=last_month
                ).values('department').distinct().count(),
            },
            
            # Performance metrics
            'avg_review_time': str(avg_review_time) if avg_review_time else None,
            'approval_rate': round(
                (approved_count / total_submissions * 100)
                if total_submissions > 0 else 0,
                1
            ),
            
            # Recent submissions
            'recent_submissions': DataSubmissionSerializer(
                base_queryset.filter(submitted_at__isnull=False)  # Only show actually submitted ones
                .order_by('-submitted_at')[:5],
                many=True
            ).data,
            
            # Template stats
            'by_template': list(
                base_queryset.values(
                    'template__code',
                    'template__name',
                    'template__criteria__number'
                )
                .annotate(
                    total=models.Count('id'),
                    pending=models.Count('id', filter=models.Q(status='submitted')),
                    approved=models.Count('id', filter=models.Q(status='approved')),
                    rejected=models.Count('id', filter=models.Q(status='rejected')),
                    draft=models.Count('id', filter=models.Q(status='draft'))
                )
                .order_by('template__criteria__number', 'template__code')
            ),
            
            # Department stats
            'by_department': list(
                base_queryset.values(
                    'department__code',
                    'department__name'
                )
                .annotate(
                    total=models.Count('id'),
                    pending=models.Count('id', filter=models.Q(status='submitted')),
                    approved=models.Count('id', filter=models.Q(status='approved')),
                    rejected=models.Count('id', filter=models.Q(status='rejected')),
                    draft=models.Count('id', filter=models.Q(status='draft'))
                )
                .order_by('department__name')
            ),
            
            # Board stats
            'by_board': list(
                base_queryset.values(
                    'template__criteria__board__code',
                    'template__criteria__board__name'
                )
                .annotate(
                    total=models.Count('id'),
                    pending=models.Count('id', filter=models.Q(status='submitted')),
                    approved=models.Count('id', filter=models.Q(status='approved')),
                    rejected=models.Count('id', filter=models.Q(status='rejected')),
                    draft=models.Count('id', filter=models.Q(status='draft')),
                    templates_count=models.Count('template', distinct=True),
                    departments_count=models.Count('department', distinct=True)
                )
                .order_by('template__criteria__board__name')
            )
        }

        return Response({
            'status': 'success',
            'data': stats
        })

        # stats['by_board'] = list(
        # base_queryset.values('board__name')
        #     .annotate(
        #         total=models.Count('id'),
        #         pending=models.Count('id', filter=models.Q(status='submitted')),
        #         approved=models.Count('id', filter=models.Q(status='approved')),
        #         rejected=models.Count('id', filter=models.Q(status='rejected'))
        #     )
        #     .order_by('-total')
        # )
        # return Response({
        #     'status': 'success',
        #     'data': stats
        # })
        
    @action(detail=False, methods=['GET'], url_path='department-breakdown', url_name='department-breakdown')
    def department_breakdown(self, request):
        logger.debug(f"Department breakdown called with academic_year: {request.query_params.get('academic_year')}")
        try:
            academic_year_id = request.query_params.get('academic_year')
            board_id = request.query_params.get('board')  # Changed to board_id for consistency

            if not board_id:
                return Response({
                    'status': 'error',
                    'message': 'Board ID is required'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get academic year
            if academic_year_id:
                academic_year = AcademicYear.objects.get(id=academic_year_id)
            else:
                academic_year = AcademicYear.objects.filter(is_current=True).first()
                
            if not academic_year:
                return Response({
                    'status': 'error',
                    'message': 'No academic year found'
                }, status=status.HTTP_404_NOT_FOUND)

            # Get board
            try:
                board = Board.objects.get(id=board_id)
            except Board.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': 'Board not found'
                }, status=status.HTTP_404_NOT_FOUND)

            # Get templates for this board only
            templates = Template.objects.filter(
                criteria__board=board
            ).select_related('criteria')
            
            total_templates = templates.count()

            if total_templates == 0:
                return Response({
                    'status': 'error',
                    'message': f'No templates found for board {board.name}'
                }, status=status.HTTP_404_NOT_FOUND)

            departments = Department.objects.all()
            department_data = []
            total_submissions = 0
            completed_submissions = 0

            for dept in departments:
                # Get submissions for this department, board and academic year
                submissions = DataSubmission.objects.filter(
                    department=dept,
                    academic_year=academic_year,
                    template__criteria__board=board
                ).select_related('template', 'verified_by')

                # Calculate department stats
                dept_completed = submissions.filter(status='approved').count()
                dept_total = total_templates
                completion_rate = (dept_completed / dept_total * 100) if dept_total > 0 else 0

                # Get template details
                template_details = []
                for template in templates:  # Iterate over board-specific templates
                    submission = submissions.filter(template=template).first()
                    template_details.append({
                        'code': template.code,
                        'name': template.name,
                        'criteria': {
                            'id': template.criteria.id,
                            'number': template.criteria.number,
                            'name': template.criteria.name
                        },
                        'status': submission.status if submission else 'pending',
                        'last_updated': submission.updated_at if submission else None,
                        'verified_by': submission.verified_by.get_full_name() if submission and submission.verified_by else None,
                        'rejection_reason': submission.rejection_reason if submission and submission.status == 'rejected' else None,
                        'submission_id': submission.id if submission else None
                    })

                department_data.append({
                    'id': dept.id,
                    'name': dept.name,
                    'code': dept.code,
                    'completion_rate': round(completion_rate, 1),
                    'completed_submissions': dept_completed,
                    'total_required': dept_total,
                    'templates': template_details
                })

                total_submissions += dept_total
                completed_submissions += dept_completed

            overall_completion_rate = (completed_submissions / total_submissions * 100) if total_submissions > 0 else 0

            return Response({
                'status': 'success',
                'data': {
                    'academic_year': {
                        'id': academic_year.id,
                        'name': academic_year.name,
                        'is_current': academic_year.is_current
                    },
                    'board': {
                        'id': board.id,
                        'name': board.name,
                        'code': board.code
                    },
                    'overall_completion_rate': round(overall_completion_rate, 1),
                    'completed_submissions': completed_submissions,
                    'total_required_submissions': total_submissions,
                    'departments': department_data
                }
            })

        except AcademicYear.DoesNotExist:
            return Response({
                'status': 'error',
                'message': 'Academic year not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error in department_breakdown: {str(e)}")
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
             
class AcademicYearTransitionViewSet(viewsets.ViewSet):
    permission_classes = [HasPermission, IsIQACDirector]

    @action(detail=True, methods=['post'])
    def start_transition(self, request, pk=None):
        """Start transition to a new academic year"""
        try:
            from_year = AcademicYear.objects.get(is_current=True)
            to_year = get_object_or_404(AcademicYear, pk=pk)

            transition_service = AcademicYearTransitionService(
                from_year=from_year,
                to_year=to_year,
                user=request.user
            )

            transition = transition_service.start_transition()
            
            # Start async task for processing
            process_academic_year_transition.delay(transition.id)

            return Response({
                'status': 'success',
                'message': 'Academic year transition started',
                'transition_id': transition.id
            })

        except ValidationError as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def transition_status(self, request, pk=None):
        """Get status of academic year transition"""
        transition = get_object_or_404(
            AcademicYearTransition,
            to_year_id=pk
        )

        return Response({
            'status': transition.status,
            'started_at': transition.started_at,
            'completed_at': transition.completed_at,
            'error_log': transition.error_log,
            'processed_by': transition.processed_by.get_full_name()
        })      

class ExportTemplateView(APIView):
    # permission_classes = [HasPermission]

    def get(self, request, *args, **kwargs):
        
        # print(request.user.roles.filter(name='IQAC Director').first())
        
        if not request.user.roles.filter(name='iqac_director').exists():
            return Response(
                {"error": "Only IQAC Director can export data"},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            # Get parameters
            academic_year_id = request.query_params.get('academic_year')
            export_type = request.query_params.get('type', 'template')
            template_code = request.query_params.get('template_code')
            criterion = request.query_params.get('criterion')
            board_code = request.query_params.get('board')

            print(f"Export parameters: year={academic_year_id}, type={export_type}, template={template_code}, criterion={criterion}, board={board_code}")

            if not board_code:
                return Response(
                    {'error': 'Board is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if not academic_year_id:
                return Response(
                    {'error': 'Academic year is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            print('_________________________________________________________')
            try:
                board = Board.objects.filter(id=board_code).first()  # Change to filter by code
                print(f"Found board: {board.name}")
                academic_year = AcademicYear.objects.get(id=academic_year_id)
            except (Board.DoesNotExist, AcademicYear.DoesNotExist) as e:
                return Response(
                    {'error': 'Invalid board or academic year'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Get templates based on export type and board
            if export_type == 'all':
                templates = Template.objects.filter(
                    criteria__board=board
                ).order_by('code')
            elif export_type == 'criterion' and criterion:
                templates = Template.objects.filter(
                    criteria__board=board,
                    code__startswith=f"{criterion}."
                ).order_by('code')
            elif export_type == 'template' and template_code:
                templates = Template.objects.filter(
                    criteria__board=board,
                    code=template_code
                )
            else:
                return Response(
                    {'error': 'Invalid export parameters'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if not templates.exists():
                return Response(
                    {'error': 'No templates found for the given criteria'},
                    status=status.HTTP_404_NOT_FOUND
                )

            print(f"Found {templates.count()} templates to export")

            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            has_data = False
            for template in templates:
                print(f"Processing template: {template.code}")
                submissions = DataSubmission.objects.filter(
                    template=template,
                    academic_year=academic_year,
                    status='approved'
                ).prefetch_related('data_rows')

                print(f"Found {submissions.count()} approved submissions for {template.code}")

                if submissions.exists():
                    has_data = True
                    ws = wb.create_sheet(title=template.code[:31])
                    exporter = ExcelExporter(template, academic_year)
                    exporter.export_to_worksheet(ws, submissions)

            if not has_data:
                return Response(
                    {'error': 'No approved submissions found'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Generate filename
            filename = f"{board.code}_{academic_year.name}"
            if export_type == 'criterion' and criterion:
                filename = f"{filename}_criterion_{criterion}"
            elif export_type == 'template' and template_code:
                filename = f"{filename}_template_{template_code}"
            filename = f"{filename}.xlsx"

            # Create response
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            print(f"Export completed successfully: {filename}")
            return response

        except Exception as e:
            print(f"Export error: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
                        
class CriteriaViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Criteria.objects.all()
    serializer_class = CriteriaSerializer
    permission_classes = [HasPermission]
    
    def get_queryset(self):
        queryset = Criteria.objects.all()
        board = self.request.query_params.get('board')
        
        if (board):
            queryset = queryset.filter(board__id=board)
            
        return queryset.order_by('order', 'number')

class BoardViewSet(APIView):
    def get(self, request):
        boards = Board.objects.all()
        return Response(BoardSerializer(boards, many=True).data)

class SettingsViewSet(viewsets.ViewSet):
    permission_classes = [HasPermission]

    @action(detail=False, methods=['get'])
    def get_settings(self, request):
        user = request.user
        settings = {
            'selectedBoard': user.profile.selected_board,
            'selectedAcademicYear': user.profile.selected_academic_year
        }
        return Response(settings)

    @action(detail=False, methods=['post'])
    def save_settings(self, request):
        user = request.user
        selected_board = request.data.get('selectedBoard')
        selected_academic_year = request.data.get('selectedAcademicYear')

        user.profile.selected_board = selected_board
        user.profile.selected_academic_year = selected_academic_year
        user.profile.save()

        return Response({'status': 'success'})
    
    

class DashboardViewSet(ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'], url_path='stats')
    def stats(self, request):
        academic_year_id = request.query_params.get('academic_year')
        board_id = request.query_params.get('board_id')
        
        stats = DashboardService.get_overall_stats(academic_year_id, board_id)
        serializer = DashboardStatsSerializer(stats)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='activity-timeline')
    def activity_timeline(self, request):
        academic_year_id = request.query_params.get('academic_year')
        board_id = request.query_params.get('board_id')
        department_id = request.query_params.get('department_id')
        days = int(request.query_params.get('days', 30))

        timeline = DashboardService.get_activity_timeline(
            academic_year_id, board_id, days, department_id
        )
        serializer = ActivityTimelineSerializer(timeline, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='criteria-completion')
    def criteria_completion(self, request):
        academic_year_id = request.query_params.get('academic_year')
        board_id = request.query_params.get('board_id')

        completion = DashboardService.get_criteria_completion(academic_year_id, board_id)
        serializer = CriteriaCompletionSerializer(completion, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='faculty-stats')
    def faculty_stats(self, request):
        academic_year_id = request.query_params.get('academic_year')
        board_id = request.query_params.get('board_id')
        department_id = request.query_params.get('department_id')

        if not department_id:
            return Response(
                {'error': 'Department ID is required'}, 
                status=400
            )

        stats = DashboardService.get_faculty_stats(
            academic_year_id, board_id, department_id
        )
        serializer = FacultyStatsSerializer(stats)
        return Response(serializer.data)